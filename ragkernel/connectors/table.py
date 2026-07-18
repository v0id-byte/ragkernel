"""工单 / 反馈表连接器（CSV / Excel）：每行 → 一条记录（一个 Page）。

ATOMIC=True 告诉 pipeline：每行是一条完整记录，整条一片、不细分，交给垂直层 classify 打分类。
SOURCE_KIND=ticket_import 供统计面板区分"导入的工单"。
"""

import csv
from pathlib import Path

from .base import Page

EXTS = {".csv", ".xlsx"}
MIME = "text/x-ticket-table"
SOURCE_KIND = "ticket_import"
ATOMIC = True


def _render_row(headers, values) -> str:
    parts = []
    for i, v in enumerate(values):
        v = "" if v is None else str(v).strip()
        if not v:
            continue
        h = str(headers[i]).strip() if i < len(headers) and headers[i] else f"列{i + 1}"
        parts.append(f"{h}：{v}")
    return "\n".join(parts)


def _read_csv(path: Path) -> list[list]:
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError:
            continue
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        return list(csv.reader(f))


def _read_xlsx(path: Path) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def load(path) -> list[Page]:
    path = Path(path)
    rows = _read_csv(path) if path.suffix.lower() == ".csv" else _read_xlsx(path)
    rows = [r for r in rows if any((c is not None and str(c).strip()) for c in r)]
    if not rows:
        return []
    headers = rows[0]
    pages = []
    for i, values in enumerate(rows[1:], start=1):
        text = _render_row(headers, values)
        if text.strip():
            pages.append(Page(text=text, page_no=i))
    return pages
